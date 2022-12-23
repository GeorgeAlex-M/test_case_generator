import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

class Config:
    def __init__(self, test_techniques, output_path):
        self.test_techniques = test_techniques
        self.output_path = output_path

class TestTechnique:
    def __init__(self, technique_type, technique, name, description, num_test_cases, coverage, inputs):
        self.type = technique_type
        self.technique = technique
        self.name = name
        self.description = description
        self.num_test_cases = num_test_cases
        self.coverage = coverage
        self.inputs = inputs

class TestCase:
    def __init__(self, test_case_num, technique, input_):
        self.test_case_num = test_case_num
        self.technique = technique
        self.input = input_

class TestCaseGenerator:
    def __init__(self, config):
        self.config = config

    def generate_test_cases(self):
        test_cases = []
        test_case_num = 1
        for technique in self.config.test_techniques:
            for i, input_ in enumerate(technique.inputs):
                test_case = TestCase(test_case_num, technique.name, input_)
                test_cases.append(test_case)
                test_case_num += 1
        return test_cases

    def write_to_excel(self):
        test_cases = self.generate_test_cases()

        wb = Workbook()
        ws = wb.active
        ws.title = "Test Cases"

        # Set the column widths
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 30

        # Freeze the first row
        ws.freeze_panes = "A2"

        # Set the cell styles
        font = Font(bold=True)
        fill = PatternFill("solid", fgColor="DDDDDD")
        align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_cells = Alignment(horizontal="left", vertical="top", wrap_text=True)
        border = Border(left=Side(border_style="thin", color="000000"),
                        right=Side(border_style="thin", color="000000"),
                        top=Side(border_style="thin", color="000000"),
                        bottom=Side(border_style="thin", color="000000"))

        # Write the header row
        ws["A1"].value = "# Test Case"
        ws["B1"].value = "Test Technique"
        ws["C1"].value = "Input"
        ws["A1"].font = font
        ws["B1"].font = font
        ws["C1"].font = font
        ws["A1"].fill = fill
        ws["B1"].fill = fill
        ws["C1"].fill = fill
        ws["A1"].alignment = align
        ws["B1"].alignment = align
        ws["C1"].alignment = align
        ws["A1"].border = border
        ws["B1"].border = border
        ws["C1"].border = border

        # Write the test case rows
        for i, test_case in enumerate(test_cases):
            ws.cell(row=i + 2, column=1, value=test_case.test_case_num)
            ws.cell(row=i + 2, column=2, value=test_case.technique)
            ws.cell(row=i + 2, column=3, value=test_case.input)
            for j in range(1, 4):
                cell = ws.cell(row=i + 2, column=j)
                cell.font = font
                cell.border = border
                cell.alignment = align_cells

        wb.save(self.config.output_path)

if __name__ == "__main__":
    test_techniques = [
        TestTechnique("dynamic", "decision table", "Decision Table", 
                      "A decision table is a way of testing all the possible combinations of input values and the expected output.",
                      2**20, "decision", ["test", "!@#$%^&*()", "1234567890", "abcdefghijklmnopqrstuvwxyz", "ABCDEFGHIJKLMNOPQRSTUVWXYZ", 
                                         "AaBbCcDdEeFfGgHhIiJjKkLlMmNnOoPpQqRrSsTtUuVvWwXxYyZz", "The quick brown fox jumps over the lazy dog.",
                                         "ąćęłńóśźżĄĆĘŁŃÓŚŹŻ", "日本語", "한국어", "Русский", "العربية", "हिन्दी", "中文", "ภาษาไทย", "Tiếng Việt", "فارسی"]),
        TestTechnique("dynamic", "boundary value analysis", "Boundary Value Analysis", 
                      "Boundary value analysis is a technique that tests the input values that are on the edges of the input domain.",
                      5, "boundary", ["", "test", "!@#$%^&*()", "1234567890", "abcdefghijklmnopqrstuvwxyz"]),
        TestTechnique("dynamic", "equivalence partitioning", "Equivalence Partitioning",
                      "Equivalence partitioning is a technique that divides the input domain into classes of input data from which test cases can be derived.",
                      3, "equivalence", ["", "abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ1234567890!@#$%^&*()", "test"]),
    TestTechnique("dynamic", "exploratory testing", "Exploratory Testing",
                  "Exploratory testing is a technique that involves testing the system with an open-minded approach, without following a predetermined test plan.",
                  10, "exploratory", ["!\"#$%&'()*+,-./:;<=>?@[\\]^_`{|}~", "admin", "password", "1234", "qwerty", "letmein", "trustno1", "sunshine", "monkey",
                                     "princess", "shadow", "master", "love", "money", "bluesky", "thunder", "superman", "dragon", "diamond", "silver", "gold",
                                     "qazqazqaz", "abcabcabc", "123123123", "!@#!@#!@#", "qwertyuiop", "asdfghjkl", "zxcvbnm", "qweqweqwe", "asdasdasd", "zxczxczxc"]),
    ]
    output_path = os.path.join(os.path.expanduser("~"), "Desktop", "test_cases.xlsx")
    config = Config(test_techniques, output_path)
    generator = TestCaseGenerator(config)
    generator.write_to_excel()
